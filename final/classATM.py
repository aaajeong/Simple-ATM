import openpyxl
import datetime
from classBankAccount import *

time = datetime.datetime.now()     #현재시간 받아옴

#엑셀파일 읽어옴
filename = "UserList.xlsx"
wb_obj = openpyxl.load_workbook(filename)
sheet_obj = wb_obj.active

max_row = 7  # 행의 개수
max_col = 3  # 열의 개수
currentCell_obj = sheet_obj.cell(row=1, column=1)
name_list = []  #이름이 들어갈 리스트
pwd_list = []    #비번이 들어갈 리스트
balance_list = [] #초기금액 들어갈 리스트

for i in range(2, max_row + 1):  # 전체 행에서
        for j in range(1, 2):  # 이름까지만
             currentCell_obj = sheet_obj.cell(row=i, column=j)
             name_list.append(currentCell_obj.value)  # 셀의 value를 Score_list에 추가한다.
        for k in range(2,3):
             currentCell_obj = sheet_obj.cell(row=i, column=k)
             pwd_list.append(currentCell_obj.value)
        for m in range(3,4):
             currentCell_obj = sheet_obj.cell(row=i, column=m)
             balance_list.append(currentCell_obj.value)

class Atm(BankAccount):
    def __init__(self, name = "None", pwd = 0, balance = 0):
        self.name = input("Enter the username: ")        #이름 입력
        if self.name in name_list:      #등록된 사용자일 경우
            print(self.name, '님 환영합니다.')
            pwd = int(input(self.name + "님의 비밀번호를 입력하세요: "))
            self.pwd = pwd
            index = name_list.index(self.name)  #입력받은 이름이 name_list의 몇번째 인덱스인지 반환
            self.balance = balance_list[index]  #해당되는 balance를 가져옴
            if (self.pwd == pwd_list[index]):   #pwd_list의 비밀번호와 비교
                print("사용자 정보가 확인되었습니다.")
                Atm.deal(self)      #거래 시작
            else:
                print("비밀번호가 틀렸습니다.")
        else:          #새로운 사용자일 경우
            register = input(self.name + "님은 등록되지 않았습니다. 추가하시겠습니까?")
            if(register == "yes"):
                Atm.reg_customer(self)  #회원 등록
                BankAccount.balance = self.balance
                Atm.deal(self)          #거래 시작
            else:
                print("거래를 종료합니다.")

    def menu(self):
        print("=" * 25)
        print("원하시는 메뉴를 선택하세요.")
        print("1. Deposit")
        print("2. Withdraw")
        print("3. Check Balance")
        print("4. Quit")
    def deal(self):
        while(1):   #4를 입력하기 전까지 무한 루프
            Atm.menu(self)
            num = input(">>")
            if(num=="1"):
                amount = int(input("입금하실 금액을 입력하세요: "))
                BankAccount.balance = self.balance
                BankAccount.deposit(BankAccount, amount)    #부모클래스의 deposit 함수를 불러온다.
                receipt = input("명세표를 출력하시겠습니까? ")
                if(receipt=="yes"):     #명세표 출력하고 싶다면
                    Atm.get_receipt1(self, amount, BankAccount.balance)     #명세표 출력
                else:
                    print("이어서 거래를 진행하세요.")
                self.balance = BankAccount.balance #처리한 결과를 저장함
            elif(num=="2"):
                amount = int(input("출금하실 금액을 입력하세요: "))
                BankAccount.balance = self.balance
                print(BankAccount.balance)
                BankAccount.withdraw(BankAccount, amount)   #부모클래스의 withdraw 함수를 불러온다.
                receipt = input("명세표를 출력하시겠습니까? ")
                if (receipt == "yes"):
                    Atm.get_receipt2(self, amount, BankAccount.balance)
                else:
                    continue
            elif(num == "3"):
                print("현재 잔액은", BankAccount.balance, "입니다.")   #앞에서 저장한 현재 잔액을 가져옴
            elif(num == "4"):
                break

    def get_receipt1(self, amount, balance):
        print("*"*25)
        print("{0: ^26}".format("명세표"))
        print("거래시간: ", time)
        print("이름: ", self.name)
        print("입금액: ", amount)  #입금한 amount를 가져옴
        print("남은 잔액: ", balance)   #입금한 뒤 잔액을 가져옴
        print("거래해주셔서 감사합니다. - by Python Bank")

    def get_receipt2(self, amount, balance):
        print("*"*25)
        print("{0:=^26}".format("명세표"))
        print("거래시간: ", time)
        print("이름: ", self.name)
        print("출금액: ", amount)
        print("남은 잔액: ",balance)
        print("거래해주셔서 감사합니다. - by Python Bank")
    def reg_customer(self):             #엑셀에 사용자 추가 하는 함수
        passwd = int(input(self.name + "님의 비밀번호를 입력하세요: "))
        new_balance = int(input(self.name + "님의 초기잔액을 입력하세요: "))
        self.balance = new_balance

        #엑셀에 사용자 추가
        max_row = 7
        newCell_obj = sheet_obj['A' + str(max_row+1)]
        newCell_obj.value = self.name
        newCell_obj = sheet_obj['B' + str(max_row+1)]
        newCell_obj.value = passwd
        newCell2_obj = sheet_obj['C' + str(max_row+1)]
        newCell2_obj.value = new_balance
        wb_obj.save("UserList.xlsx")
        print("등록이 완료되었습니다!")
        max_row = max_row+1


















